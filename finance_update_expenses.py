from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
from array import *
import robin_stocks.robinhood as rob
import datetime

rob.login(username="jacobramsey31@gmail.com",
         password="F(x)=x^2+3x",
         expiresIn=86400,
         by_sms=True)  

expensises = load_workbook(r'c:\Users\jacob\OneDrive\Documents\expensises.xlsx')
month_yr=expensises["Sheet3"].cell(3,3).value
prof_dict=rob.profiles.load_account_profile(info=None)
port_dict=rob.profiles.load_portfolio_profile(info=None)



stockDict=rob.get_all_positions()
date1=datetime.datetime.now()
date2=date1.strftime("%Y-%m")

i=3
dict1={}
while expensises["Sheet3"].cell(i,3).value!=None:
    temp1=expensises["Sheet3"].cell(i,3).value
    if temp1.strftime("%Y-%m") == date2 :
        array = array('f', [float(prof_dict["portfolio_cash"])+float(port_dict["extended_hours_market_value"]),float(port_dict["equity"])-float(expensises["Sheet3"].cell(i-1,5).value)])
        if expensises["Sheet3"].cell(i,8).value==None:
            expensises["Sheet3"].cell(row=i, column=8, value =prof_dict["portfolio_cash"])
        expensises["Sheet3"].cell(row=i, column=5, value =array[0])
        
        placeHolder=2
        TotMonStockSpent=0
        spendingVal=float(expensises["Sheet3"].cell(i,6).value)
        
        if float(expensises["Sheet3"].cell(i,8).value)> float(prof_dict["portfolio_cash"]):
            spendingVal-=float(prof_dict["portfolio_cash"])-float(expensises["Sheet3"].cell(i,8).value)
            
        p=2
        for h in stockDict:
            #print(spendingVal)
            if h['quantity']!='0.00000000':
                #print(h)
                g=h['updated_at'].find('T')
                dt=h['updated_at'][0:g]
                dt1=datetime.datetime.strptime(dt,"%Y-%m-%d")
                dt1=dt1.strftime("%Y-%m-%d")
                #print(dt1)
                
                if dt1 == date1.strftime("%Y-%m-%d"):
                    
                    change=float(expensises["Sheet3"].cell(p,11).value)*float(expensises["Sheet3"].cell(p,12).value)-float(h["quantity"])*float(h["average_buy_price"])
                    
                    spendingVal-=change
                    
                    TotMonStockSpent+=change
                    p+=1
                expensises["Sheet3"].cell(row=placeHolder, column=10, value =h["instrument"])
                expensises["Sheet3"].cell(row=placeHolder, column=11, value =h["average_buy_price"])
                expensises["Sheet3"].cell(row=placeHolder, column=12, value =h["quantity"])
                placeHolder+=1
        expensises["Sheet3"].cell(row=i, column=6, value =spendingVal)
        expensises["Sheet3"].cell(row=i, column=7, value =port_dict["extended_hours_market_value"])
        expensises["Sheet3"].cell(row=i, column=8, value =prof_dict["portfolio_cash"])
        expensises["Sheet3"].cell(row=i, column=9, value =date1)
    i+=1
expensises.save(r'c:\Users\jacob\OneDrive\Documents\expensises.xlsx')
rob.logout()